import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models import SteelWip, QrCodes, Locations


# ── 고정 컬럼값 (import 파일 형식과 동일) ─────────────────────────────────────

FIXED_VALUES = {
    "재고사업장":  "(주)유석철강 청주지점",
    "관리사업장":  "(주)유석철강 청주지점",
    "품목계정":    "재공품",
    "등급":        "정상품",
    "품명":        "후판",
    "도료":        "없음",
    "예약수량":    0,
    "예약중량":    0,
    "재고구분":    "자사",
    "창고":        "Laser 창고(오창)",
    "적재위치":    "공통",
    "소유주":      "유석철강(청주지점)",
    "용도":        "없음",
    "상태":        "정상",
}

# import 파일과 동일한 컬럼 순서
COLUMNS = [
    "재고사업장", "관리사업장", "품목계정", "등급", "품명", "도료",
    "소재번호",   "재질",       "두께",     "폭",   "길이",
    "재고중량",   "예약수량",   "예약중량",
    "재고구분",   "창고",       "적재위치",
    "위치",       "소유주",     "용도",     "상태",
    "제조사",
]


# ── 데이터 조회 ────────────────────────────────────────────────────────────────

async def _fetch_rows(db: AsyncSession) -> list[dict]:
    wip_result = await db.execute(select(SteelWip))
    wips = wip_result.scalars().all()

    if not wips:
        return []

    # qr_id → qr_code 맵
    qr_result = await db.execute(select(QrCodes))
    qr_map: dict[int, str] = {
        r.id: r.qr_code for r in qr_result.scalars().all() if r.qr_code
    }

    # location_id → loc_name 맵
    loc_result = await db.execute(select(Locations))
    loc_map: dict[int, str] = {
        r.id: r.loc_name for r in loc_result.scalars().all() if r.loc_name
    }

    rows = []
    for wip in wips:
        qr_code = qr_map.get(wip.qr_id, "") if wip.qr_id else ""

        # 위치: loc_name-stack_level
        if wip.location_id and wip.stack_level is not None:
            loc_name = loc_map.get(wip.location_id, "")
            location_str = f"{loc_name}-{wip.stack_level}" if loc_name else ""
        else:
            location_str = ""

        row = {col: FIXED_VALUES.get(col, "") for col in COLUMNS}
        row.update({
            "소재번호":  qr_code,
            "재질":      wip.material or "",
            "두께":      wip.thickness,
            "폭":        wip.width,
            "길이":      wip.length,
            "재고중량":  wip.weight,
            "위치":      location_str,
            "제조사":    wip.manufacturer or "",
        })
        rows.append(row)

    return rows


# ── CSV 생성 ──────────────────────────────────────────────────────────────────

async def export_wip_csv(db: AsyncSession) -> bytes:
    rows = await _fetch_rows(db)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS)
    writer.writeheader()
    writer.writerows(rows)

    # cp949로 인코딩 (한글 Excel 호환)
    return buf.getvalue().encode("utf-8-sig")


# ── XLSX 생성 ─────────────────────────────────────────────────────────────────

async def export_wip_xlsx(db: AsyncSession) -> bytes:
    rows = await _fetch_rows(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "재고현황"

    # 헤더 스타일
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # 데이터 행
    body_font = Font(name="맑은 고딕", size=10)
    body_align = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.font = body_font
            cell.alignment = body_align

    # 열 너비 자동 조정 (최소 8, 최대 30)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(row.get(col_name, "") or "")) for row in rows) if rows else [0],
        )
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = min(max(max_len + 2, 8), 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()